import os, json, re, io, base64, urllib.request, urllib.parse
from http.server import BaseHTTPRequestHandler
from openpyxl import load_workbook

SUPABASE_URL = os.environ.get('SUPABASE_URL', 'https://qlvlpvgyjoeprvghmoyn.supabase.co')
SUPABASE_KEY = os.environ.get('SUPABASE_SERVICE_ROLE_KEY', '')
TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), 'template.xlsx')

SHEET_SECTION_MAP = {
    'Governance': ' Governance',
    'Human Rights and Labour': 'Human Rights & Labour',
    'Environment': 'Environment',
    'Anti-Corruption': ' Anti-Corruption',
    'CEO Statement': 'CEO Statement',
}

Q_PATTERN = re.compile(
    r'^(G\d+(?:\.\d+)?|HR/L\d+(?:\.\d+)?|E\d+(?:\.\d+)?|AC\d+(?:\.\d+)?'
    r'|C\d+|S\d+|R\d+)[.\s]'
)
# Suffix pattern: letters after the numeric part (e.g., "2A", "3AA", "9B")
SUFFIX_PATTERN = re.compile(r'^(.*?\d+)([A-Z]+)$')

CHECKBOX = '❑'
CHECKED  = '☑'
RADIO    = '🔾'
SELECTED = '🔘'


def supabase_get(path):
    req = urllib.request.Request(
        f"{SUPABASE_URL}/rest/v1/{path}",
        headers={
            'apikey': SUPABASE_KEY,
            'Authorization': f'Bearer {SUPABASE_KEY}',
        }
    )
    with urllib.request.urlopen(req, timeout=20) as resp:
        return json.loads(resp.read())


def normalize(s):
    """Collapse whitespace and lowercase for fuzzy comparison."""
    return re.sub(r'\s+', ' ', str(s or '').lower().strip()
                  .replace('\r\n', ' ').replace('\n', ' ').replace('\r', ' '))


def texts_match(a, b):
    a, b = normalize(a), normalize(b)
    if not a or not b:
        return False
    return a == b or b.startswith(a) or a.startswith(b)


def parse_question_locs(ws):
    """Return ordered list of (qid, row_idx_0based) from the sheet."""
    locs = []
    for i, row in enumerate(ws.iter_rows()):
        val = row[1].value if len(row) > 1 else None
        if val:
            m = Q_PATTERN.match(str(val).strip())
            if m:
                locs.append((m.group(1), i))
    return locs


def get_question_text(ws, row_idx):
    """Return the question text (column B) at row_idx."""
    for i, row in enumerate(ws.iter_rows()):
        if i == row_idx:
            return str(row[1].value or '').strip()
    return ''


def fill_sheet(ws, subs):
    """
    Fill one worksheet with the given submissions.
    Returns (filled_qids: list, unfilled_qids: list)
    """
    all_rows = list(ws.iter_rows())
    n = len(all_rows)

    q_locs = parse_question_locs(ws)
    q_dict = {qid: idx for qid, idx in q_locs}

    filled_qids = []
    unfilled_subs = []

    for sub in subs:
        qid = sub['question_id']
        choice   = str(sub.get('choice', '')   or '').strip()
        subq     = str(sub.get('subquestion', '') or '').strip()
        response = str(sub.get('response', '')  or '').strip()

        # Resolve question row: try exact, then strip letter suffix
        q_row = q_dict.get(qid)
        parent_qid = qid
        if q_row is None:
            sm = SUFFIX_PATTERN.match(qid)
            if sm:
                parent_qid = sm.group(1)
                q_row = q_dict.get(parent_qid)

        if q_row is None:
            unfilled_subs.append(sub)
            continue

        # Range for this question (up to start of next question)
        end_row = n
        for _, other_row in q_locs:
            if other_row > q_row and other_row < end_row:
                end_row = other_row

        success = False

        # ── Text response (e.g. AC2A, G6.1A) ──────────────────────────────
        if response and not choice:
            for i in range(q_row, end_row):
                cell = all_rows[i][1] if len(all_rows[i]) > 1 else None
                if cell and cell.value and 'Please provide additional information' in str(cell.value):
                    cell.value = 'Response: ' + response
                    success = True
                    break

        # ── Checkbox question (❑ rows, no subquestion) ─────────────────────
        elif choice and not subq:
            for i in range(q_row, end_row):
                cell = all_rows[i][1] if len(all_rows[i]) > 1 else None
                if cell and cell.value and str(cell.value).startswith(CHECKBOX):
                    option_text = str(cell.value)[1:].strip()
                    if texts_match(choice, option_text):
                        cell.value = CHECKED + ' ' + option_text
                        success = True
                        # no break — multiple checkboxes may be selected

        # ── Matrix question (subquestion + choice) ─────────────────────────
        elif choice and subq:
            header_row = None
            header_cols = {}
            for i in range(q_row + 1, end_row):
                row = all_rows[i]
                b_val = row[1].value if len(row) > 1 else None
                c_val = row[2].value if len(row) > 2 else None
                if (not b_val or str(b_val).strip() == '') and c_val:
                    header_row = i
                    for j, cell in enumerate(row):
                        if j >= 2 and cell.value:
                            header_cols[j] = str(cell.value)
                    break

            if header_row is not None:
                target_col = None
                for col_idx, header_text in header_cols.items():
                    if texts_match(choice, header_text):
                        target_col = col_idx
                        break

                if target_col is not None:
                    for i in range(header_row + 1, end_row):
                        row = all_rows[i]
                        label = row[1].value if len(row) > 1 else None
                        if label and texts_match(subq, str(label)):
                            if len(row) > target_col:
                                cell = row[target_col]
                                if str(cell.value or '').strip() == RADIO:
                                    cell.value = SELECTED
                                    success = True
                            break

        if success:
            filled_qids.append(parent_qid if parent_qid != qid else qid)
        else:
            unfilled_subs.append(sub)

    return filled_qids, unfilled_subs


def get_all_excel_qids(wb):
    """Return all question IDs and their text from all answer sheets."""
    results = []
    for section, sheet_name in SHEET_SECTION_MAP.items():
        try:
            ws = wb[sheet_name]
        except KeyError:
            continue
        for qid, row_idx in parse_question_locs(ws):
            text = get_question_text(ws, row_idx)
            results.append({'qid': qid, 'section': section, 'text': text})
    return results


class handler(BaseHTTPRequestHandler):
    def do_OPTIONS(self):
        self.send_response(200)
        self._cors()
        self.end_headers()

    def do_POST(self):
        try:
            length = int(self.headers.get('Content-Length', 0))
            body = json.loads(self.rfile.read(length) or b'{}')
            company_name = str(body.get('company_name', '')).strip()
            if not company_name:
                return self._error('company_name is required')

            # ── Fetch submissions ──────────────────────────────────────────
            encoded = urllib.parse.quote(company_name)
            subs = supabase_get(
                f'company_submissions?network=eq.MY'
                f'&company_name=eq.{encoded}'
                f'&select=question_id,section,subquestion,choice,response'
                f'&order=section,question_id'
            )

            # Deduplicate
            seen = set()
            unique_subs = []
            for s in subs:
                key = (s['question_id'], s.get('subquestion', ''),
                       s.get('choice', ''), s.get('response', ''))
                if key not in seen:
                    seen.add(key)
                    unique_subs.append(s)

            # ── Group by section ───────────────────────────────────────────
            by_section = {}
            for s in unique_subs:
                sec = s.get('section', '')
                by_section.setdefault(sec, []).append(s)

            # ── Load and fill template ─────────────────────────────────────
            wb = load_workbook(TEMPLATE_PATH)

            all_filled = []
            for section, subs_in_section in by_section.items():
                sheet_name = SHEET_SECTION_MAP.get(section)
                if not sheet_name:
                    continue
                try:
                    ws = wb[sheet_name]
                except KeyError:
                    continue
                filled, _ = fill_sheet(ws, subs_in_section)
                all_filled.extend(filled)

            # ── Determine pending questions ────────────────────────────────
            all_excel_qids = get_all_excel_qids(wb)
            filled_set = set(all_filled)
            pending = [
                q for q in all_excel_qids
                if q['qid'] not in filled_set
            ]

            # ── Serialize to bytes ─────────────────────────────────────────
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            excel_b64 = base64.b64encode(buf.read()).decode()

            safe_name = re.sub(r'[^\w\s-]', '', company_name).strip().replace(' ', '_')
            filename = f"CoP_2026_{safe_name}.xlsx"

            response_body = json.dumps({
                'excel_base64': excel_b64,
                'filename': filename,
                'total_filled': len(set(all_filled)),
                'total_questions': len(all_excel_qids),
                'pending': pending,
            }).encode()

            self.send_response(200)
            self._cors()
            self.send_header('Content-Type', 'application/json')
            self.send_header('Content-Length', str(len(response_body)))
            self.end_headers()
            self.wfile.write(response_body)

        except Exception as e:
            self._error(str(e))

    def _cors(self):
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')

    def _error(self, msg):
        body = json.dumps({'error': msg}).encode()
        self.send_response(500)
        self._cors()
        self.send_header('Content-Type', 'application/json')
        self.send_header('Content-Length', str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def log_message(self, *args):
        pass
